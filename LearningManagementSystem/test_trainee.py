import unittest
import openpyxl
from modules.trainee import Trainee


class TestTrainee(unittest.TestCase):
    def setUp(self):
        self.t1 = Trainee('1', 'Irina', 'Python', 'Bachelor', '1 year')
        self.t2 = Trainee('2', 'Bob', 'Java', 'Master', '2 years')
        self.t3 = Trainee('3', 'Charles', 'C++', 'PhD', '3 years')

    def test_save_to_excel(self):
        self.t1.save_to_excel()
        self.t2.save_to_excel()
        self.t3.save_to_excel()

        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainees']
        self.assertEqual(ws.cell(row=2, column=1).value, '1')
        self.assertEqual(ws.cell(row=2, column=2).value, 'Irina')
        self.assertEqual(ws.cell(row=2, column=3).value, 'Python')
        self.assertEqual(ws.cell(row=2, column=4).value, 'Bachelor')
        self.assertEqual(ws.cell(row=2, column=5).value, '1 year')
        self.assertEqual(ws.cell(row=3, column=1).value, '2')
        self.assertEqual(ws.cell(row=3, column=2).value, 'Bob')
        self.assertEqual(ws.cell(row=3, column=3).value, 'Java')
        self.assertEqual(ws.cell(row=3, column=4).value, 'Master')
        self.assertEqual(ws.cell(row=3, column=5).value, '2 years')
        self.assertEqual(ws.cell(row=4, column=1).value, '3')
        self.assertEqual(ws.cell(row=4, column=2).value, 'Charles')
        self.assertEqual(ws.cell(row=4, column=3).value, 'C++')
        self.assertEqual(ws.cell(row=4, column=4).value, 'PhD')
        self.assertEqual(ws.cell(row=4, column=5).value, '3 years')

    def test_delete_from_excel(self):
        self.t1.delete_from_excel()

        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainees']
        self.assertIsNone(ws.cell(row=2, column=1).value)
        self.assertIsNone(ws.cell(row=2, column=2).value)
        self.assertIsNone(ws.cell(row=2, column=3).value)
        self.assertIsNone(ws.cell(row=2, column=4).value)
        self.assertIsNone(ws.cell(row=2, column=5).value)

    def test_update_in_excel(self):
        self.t2.set_name('Bobby')
        self.t2.set_course('C#')
        self.t2.set_degree('PhD')
        self.t2.set_work_exp('4 years')
        self.t2.update_in_excel()

        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainees']
        self.assertEqual(ws.cell(row=3, column=1).value, '2')
        self.assertEqual(ws.cell(row=3, column=2).value, 'Bobby')
        self.assertEqual(ws.cell(row=3, column=3).value, 'C#')
        self.assertEqual(ws.cell(row=3, column=4).value, 'PhD')
        self.assertEqual(ws.cell(row=3, column=5).value, '4 years')


if __name__ == '__main__':
    unittest.main()
