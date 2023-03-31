import unittest
from modules.courses import Course
import openpyxl


class TestCourse(unittest.TestCase):
    def setUp(self):
        self.test_course = Course("C001", "Introduction to Python")

    def test_get_id(self):
        self.assertEqual(self.test_course.get_id(), "C001")

    def test_set_id(self):
        self.test_course.set_id("C002")
        self.assertEqual(self.test_course.get_id(), "C002")

    def test_get_description(self):
        self.assertEqual(self.test_course.get_description(), "Introduction to Python")

    def test_set_description(self):
        self.test_course.set_description("Advanced Python")
        self.assertEqual(self.test_course.get_description(), "Advanced Python")

    def test_save_to_excel(self):
        self.test_course.save_to_excel()
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfCourses']
        cell_value = ws.cell(row=ws.max_row, column=1).value
        self.assertEqual(cell_value, "C002")
        wb.close()

    def test_delete_from_excel(self):
        self.test_course.delete_from_excel()
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfCourses']
        cell_found = False
        for row in ws.iter_rows(min_row=2, max_col=1):
            for cell in row:
                if cell.value == "C002":
                    cell_found = True
        self.assertFalse(cell_found)
        wb.close()

    def test_update_in_excel(self):
        self.test_course.set_description("Intermediate Python")
        self.test_course.update_in_excel()
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfCourses']
        cell_value = ws.cell(row=2, column=2).value
        self.assertEqual(cell_value, "Intermediate Python")
        wb.close()


if __name__ == '__main__':
    unittest.main()
