import unittest

import openpyxl

from modules.manager import Manager


class TestManager(unittest.TestCase):
    def test_save_to_excel(self):
        m = Manager('M001', 'Tomi Janet', 'tomi@tomijanet.com', '1234567890', ['C001'], ['T001'])
        m.save_to_excel()
        # Check if the manager record was added to the Excel file
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['Managers']
        trainer_id = ws.cell(row=ws.max_row, column=1).value
        full_name = ws.cell(row=ws.max_row, column=2).value
        email = ws.cell(row=ws.max_row, column=3).value
        phone_number = ws.cell(row=ws.max_row, column=4).value
        managed_courses = ws.cell(row=ws.max_row, column=5).value.split(',')
        managed_trainers = ws.cell(row=ws.max_row, column=6).value.split(',')
        self.assertEqual(trainer_id, 'M001')
        self.assertEqual(full_name, 'Tomi Janet')
        self.assertEqual(email, 'tomi@tomijanet.com')
        self.assertEqual(phone_number, '1234567890')
        self.assertEqual(managed_courses, ['C001'])
        self.assertEqual(managed_trainers, ['T001'])

    def test_update_in_excel(self):
        m = Manager('M001', 'Tomi Janet', 'tomi@tomijanet.com', '1234567890', ['C001'], ['T001'])
        m.update_in_excel(new_full_name='Jane Doe', new_managed_courses=['C001', 'C002'])
        # Check if the manager record was updated in the Excel file
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['Managers']
        full_name = None
        managed_courses = None
        for row in ws.iter_rows(min_row=2, max_col=6):
            if row[0].value == 'M001':
                full_name = row[1].value
                managed_courses = row[4].value.split(',')
                break
        self.assertEqual(full_name, 'Jane Doe')
        self.assertEqual(managed_courses, ['C001', 'C002'])

    def test_delete_from_excel(self):
        m = Manager('M001', 'Tomi Janet', 'tomi@tomijanet.com', '1234567890', ['C001', 'C002'], ['T001', 'T002'])
        m.save_to_excel()
        m.delete_from_excel()
        # Check if the manager record was deleted from the Excel file
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['Managers']
        full_name = None
        for row in ws.iter_rows(min_row=2, max_col=6):
            if row[0].value == 'M001':
                full_name = row[1].value
                break
        self.assertIsNone(full_name)


if __name__ == '__main__':
    unittest.main()
