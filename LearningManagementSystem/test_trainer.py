import unittest
from modules.trainer import Trainer
import openpyxl


class TestTrainer(unittest.TestCase):
    def setUp(self):
        self.test_trainer = Trainer("T001", "Shilpa Vaidya", "shilpa@shilpa.com", "123-456-7890")

    def test_get_trainer_id(self):
        self.assertEqual(self.test_trainer.get_trainer_id(), "T001")

    def test_set_trainer_id(self):
        self.test_trainer.set_trainer_id("T002")
        self.assertEqual(self.test_trainer.get_trainer_id(), "T002")

    def test_get_full_name(self):
        self.assertEqual(self.test_trainer.get_full_name(), "Shilpa Vaidya")

    def test_set_full_name(self):
        self.test_trainer.set_full_name("Tom Ford")
        self.assertEqual(self.test_trainer.get_full_name(), "Tom Ford")

    def test_get_email(self):
        self.assertEqual(self.test_trainer.get_email(), "shilpa@shilpa.com")

    def test_set_email(self):
        self.test_trainer.set_email("tom@tomford.com")
        self.assertEqual(self.test_trainer.get_email(), "tom@tomford.com")

    def test_get_phone_number(self):
        self.assertEqual(self.test_trainer.get_phone_number(), "123-456-7890")

    def test_set_phone_number(self):
        self.test_trainer.set_phone_number("987-654-3210")
        self.assertEqual(self.test_trainer.get_phone_number(), "987-654-3210")

    def test_save_to_excel(self):
        self.test_trainer.save_to_excel()
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainers']
        cell_value = ws.cell(row=ws.max_row, column=1).value
        self.assertEqual(cell_value, "T001")
        wb.close()

    def test_delete_from_excel(self):
        self.test_trainer.delete_from_excel()
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainers']
        cell_found = False
        for row in ws.iter_rows(min_row=2, max_col=1):
            for cell in row:
                if cell.value == "T002":
                    cell_found = True
        self.assertFalse(cell_found)
        wb.close()

    def test_update_in_excel(self):
        self.test_trainer.set_full_name("Tom Ford")
        self.test_trainer.update_in_excel()
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['ListOfTrainers']
        cell_value = ws.cell(row=2, column=2).value
        self.assertEqual(cell_value, "Tom Ford")
        wb.close()


if __name__ == '__main__':
    unittest.main()
