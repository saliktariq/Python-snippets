import unittest
import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter
from modules.attendance import Attendance


class TestAttendance(unittest.TestCase):
    def test_constructor(self):
        att = Attendance(start_time='09:00', end_time='17:00', trainer_id='TR001', trainer_name='Shilpa Vaidya')
        self.assertIsInstance(att, Attendance)
        self.assertEqual(att.start_time, '09:00')
        self.assertEqual(att.end_time, '17:00')
        self.assertEqual(att.trainer_id, 'TR001')
        self.assertEqual(att.trainer_name, 'Shilpa Vaidya')
        self.assertEqual(att.trainees_attended, [])
        self.assertEqual(att.trainees_not_attended, [])

    def test_save_to_excel(self):
        # Create an instance of the Attendance class
        att = Attendance(start_time='09:00', end_time='17:00', trainer_id='TR001', trainer_name='Shilpa Vaidya')
        att.trainees_attended = [('TRN001', 'Alice'), ('TRN002', 'Bob')]
        att.trainees_not_attended = [('TRN003', 'Charlie'), ('TRN004', 'David')]
        # Call the save_to_excel method
        att.save_to_excel()
        # Check that the data was saved to the correct sheet and cell
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        sheet_name = datetime.today().strftime('%m-%d-%Y')
        ws = wb[sheet_name]
        self.assertEqual(ws['A2'].value, 'TRN001')
        self.assertEqual(ws['B2'].value, 'Alice')
        self.assertEqual(ws['C2'].value, 'Present')
        self.assertEqual(ws['A3'].value, 'TRN002')
        self.assertEqual(ws['B3'].value, 'Bob')
        self.assertEqual(ws['C3'].value, 'Present')
        self.assertEqual(ws['A4'].value, 'TRN003')
        self.assertEqual(ws['B4'].value, 'Charlie')
        self.assertEqual(ws['C4'].value, 'Absent')
        self.assertEqual(ws['A5'].value, 'TRN004')
        self.assertEqual(ws['B5'].value, 'David')
        self.assertEqual(ws['C5'].value, 'Absent')
        self.assertEqual(ws['D1'].value, 'Start Time')
        self.assertEqual(ws['E1'].value, '09:00')
        self.assertEqual(ws['D2'].value, 'End Time')
        self.assertEqual(ws['E2'].value, '17:00')
        self.assertEqual(ws['D3'].value, 'Trainer ID')
        self.assertEqual(ws['E3'].value, 'TR001')
        self.assertEqual(ws['D4'].value, 'Trainer Name')
        self.assertEqual(ws['E4'].value, 'Shilpa Vaidya')
        # Clean up by removing the sheet we created
        wb.remove(ws)
        wb.save('MasterRecord.xlsx')
