import openpyxl


class Enrollment:
    def __init__(self, course_id, trainee_id):
        self._course_id = course_id
        self._trainee_id = trainee_id

    def get_course_id(self):
        return self._course_id

    def set_course_id(self, course_id):
        self._course_id = course_id

    def get_trainee_id(self):
        return self._trainee_id

    def set_trainee_id(self, trainee_id):
        self._trainee_id = trainee_id

    def save_to_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['Enrollments']
            last_row = ws.max_row
            ws.cell(row=last_row + 1, column=1, value=self._course_id)
            ws.cell(row=last_row + 1, column=2, value=self._trainee_id)
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error saving enrollment to Excel file: {e}")

    def update_in_excel(self, new_course_id=None, new_trainee_id=None):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['Enrollments']
            for row in ws.iter_rows(min_row=2, max_col=2):
                if row[0].value == self._course_id and row[1].value == self._trainee_id:
                    if new_course_id is not None:
                        row[0].value = new_course_id
                    if new_trainee_id is not None:
                        row[1].value = new_trainee_id
                    break
            else:
                print("Enrollment not found.")
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error updating enrollment in Excel file: {e}")

    def delete_from_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['Enrollments']
            for row in ws.iter_rows(min_row=2, max_col=2):
                if row[0].value == self._course_id and row[1].value == self._trainee_id:
                    ws.delete_rows(row[0].row, 1)
                    break
            else:
                print("Enrollment not found.")
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error deleting enrollment from Excel file: {e}")
