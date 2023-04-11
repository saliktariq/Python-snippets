import openpyxl


class Course:
    def __init__(self, course_id, description):
        self._id = course_id
        self._description = description

    def get_id(self):
        return self._id

    def set_id(self, course_id):
        self._id = course_id

    def get_description(self):
        return self._description

    def set_description(self, description):
        self._description = description

    def save_to_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfCourses']
            last_row = ws.max_row
            ws.cell(row=last_row + 1, column=1, value=self._id)
            ws.cell(row=last_row + 1, column=2, value=self._description)
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error saving course to Excel file: {e}")

    def delete_from_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfCourses']
            for row in ws.iter_rows(min_row=2, max_col=1):
                for cell in row:
                    if cell.value == self._id:
                        ws.delete_rows(cell.row, 1)
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error deleting course from Excel file: {e}")

    def update_in_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfCourses']
            for row in ws.iter_rows(min_row=2, max_col=2):
                if row[0].value == self._id:
                    row[1].value = self._description
                    break
            else:
                print("Course not found.")
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error updating course in Excel file: {e}")

    @staticmethod
    def get_all_from_excel():
        courses = []
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
        except FileNotFoundError:
            return courses
        ws = wb['ListOfCourses'] if 'ListOfCourses' in wb.sheetnames else None
        if ws:
            for row in ws.iter_rows(min_row=2, values_only=True):
                course = Course(row[0], row[1])
                courses.append(course)
        return courses
