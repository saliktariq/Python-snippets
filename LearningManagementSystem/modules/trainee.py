import openpyxl


class Trainee:
    def __init__(self, trainee_id, name, course, degree, work_exp):
        self._id = trainee_id
        self._name = name
        self._course = course
        self._degree = degree
        self._work_exp = work_exp

    def get_id(self):
        return self._id

    def set_id(self, trainee_id):
        self._id = trainee_id

    def get_name(self):
        return self._name

    def set_name(self, name):
        self._name = name

    def get_course(self):
        return self._course

    def set_course(self, course):
        self._course = course

    def get_degree(self):
        return self._degree

    def set_degree(self, degree):
        self._degree = degree

    def get_work_exp(self):
        return self._work_exp

    def set_work_exp(self, work_exp):
        self._work_exp = work_exp

    def save_to_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfTrainees']
            last_row = ws.max_row
            ws.cell(row=last_row + 1, column=1, value=self._id)
            ws.cell(row=last_row + 1, column=2, value=self._name)
            ws.cell(row=last_row + 1, column=3, value=self._course)
            ws.cell(row=last_row + 1, column=4, value=self._degree)
            ws.cell(row=last_row + 1, column=5, value=self._work_exp)
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error saving trainee to Excel file: {e}")

    def delete_from_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfTrainees']
            for row in ws.iter_rows(min_row=2, max_col=1):
                for cell in row:
                    if cell.value == self._id:
                        ws.delete_rows(cell.row, 1)
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error deleting trainee from Excel file: {e}")

    def update_in_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfTrainees']
            for row in ws.iter_rows(min_row=2, max_col=5):
                if row[0].value == self._id:
                    row[1].value = self._name
                    row[2].value = self._course
                    row[3].value = self._degree
                    row[4].value = self._work_exp
                    break
            else:
                print("Trainee not found.")
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error updating trainee in Excel file: {e}")
