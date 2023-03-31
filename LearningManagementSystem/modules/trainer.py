import openpyxl


class Trainer:
    def __init__(self, trainer_id, full_name, email, phone_number):
        self._trainer_id = trainer_id
        self._full_name = full_name
        self._email = email
        self._phone_number = phone_number

    def get_trainer_id(self):
        return self._trainer_id

    def set_trainer_id(self, trainer_id):
        self._trainer_id = trainer_id

    def get_full_name(self):
        return self._full_name

    def set_full_name(self, full_name):
        self._full_name = full_name

    def get_email(self):
        return self._email

    def set_email(self, email):
        self._email = email

    def get_phone_number(self):
        return self._phone_number

    def set_phone_number(self, phone_number):
        self._phone_number = phone_number

    def save_to_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfTrainers']
            last_row = ws.max_row
            ws.cell(row=last_row + 1, column=1, value=self._trainer_id)
            ws.cell(row=last_row + 1, column=2, value=self._full_name)
            ws.cell(row=last_row + 1, column=3, value=self._email)
            ws.cell(row=last_row + 1, column=4, value=self._phone_number)
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error saving trainer to Excel file: {e}")

    def delete_from_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfTrainers']
            for row in ws.iter_rows(min_row=2, max_col=1):
                for cell in row:
                    if cell.value == self._trainer_id:
                        ws.delete_rows(cell.row, 1)
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error deleting trainer from Excel file: {e}")

    def update_in_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['ListOfTrainers']
            for row in ws.iter_rows(min_row=2, max_col=4):
                if row[0].value == self._trainer_id:
                    row[1].value = self._full_name
                    row[2].value = self._email
                    row[3].value = self._phone_number
                    break
            else:
                print("Trainer not found.")
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error updating trainer in Excel file: {e}")
