import os
import smtplib
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart

import openpyxl
from openpyxl.utils import get_column_letter


class Attendance:
    def __init__(self, start_time, end_time, trainer_id, trainer_name, trainees_attended=None, trainees_not_attended=None):
        self.date = datetime.today().strftime('%m-%d-%Y')
        self.start_time = start_time
        self.end_time = end_time
        self.trainer_id = trainer_id
        self.trainer_name = trainer_name
        if trainees_attended is None:
            self.trainees_attended = []
        else:
            self.trainees_attended = trainees_attended
        if trainees_not_attended is None:
            self.trainees_not_attended = []
        else:
            self.trainees_not_attended = trainees_not_attended

    def save_to_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
        except FileNotFoundError:
            wb = openpyxl.Workbook()
        sheet_name = self.date
        # Check if the sheet already exists
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            row_count = ws.max_row
        # Create a new sheet if it doesn't exist
        else:
            ws = wb.create_sheet(title=sheet_name)
            row_count = 1
            # Write headers to the first row of the sheet
            headers = ['Trainee ID', 'Trainee Name', 'Attendance']
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                cell = ws['{}{}'.format(col_letter, row_count)]
                cell.value = header
        # Write attendance data to the sheet
        row_count += 1
        for trainee_id, trainee_name in self.trainees_attended:
            ws.cell(row=row_count, column=1, value=trainee_id)
            ws.cell(row=row_count, column=2, value=trainee_name)
            ws.cell(row=row_count, column=3, value='Present')
            row_count += 1
        for trainee_id, trainee_name in self.trainees_not_attended:
            ws.cell(row=row_count, column=1, value=trainee_id)
            ws.cell(row=row_count, column=2, value=trainee_name)
            ws.cell(row=row_count, column=3, value='Absent')
            row_count += 1
        # Write metadata to the sheet
        metadata = {
            'Start Time': self.start_time,
            'End Time': self.end_time,
            'Trainer ID': self.trainer_id,
            'Trainer Name': self.trainer_name,
        }
        for col_num, (key, value) in enumerate(metadata.items(), 1):
            col_letter = get_column_letter(ws.max_column + 1)
            cell = ws['{}{}'.format(col_letter, 1)]
            cell.value = key
            cell = ws['{}{}'.format(col_letter, 2)]
            cell.value = value
        # Save the workbook
        wb.save('MasterRecord.xlsx')


    @staticmethod
    def send_attendance_email(to_addr):
        # Get the filename of the saved attendance sheet
        filename = datetime.today().strftime('%m-%d-%Y') + '.xlsx'

        # Open the saved attendance sheet
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb[datetime.today().strftime('%m-%d-%Y')]

        # Create a MIME multipart message
        msg = MIMEMultipart()
        msg['From'] = 'sender@emailaddressgmail.com'
        msg['To'] = to_addr
        msg['Subject'] = 'Attendance Sheet'

        # Attach the attendance sheet to the email as an
        # application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
        with open(filename, 'wb') as f:
            wb.save(f)
        with open(filename, 'rb') as f:
            attachment = MIMEApplication(f.read(), _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            attachment.add_header('Content-Disposition', 'attachment', filename=filename)
            msg.attach(attachment)

        # Send the email
        smtp_server = 'smtp server address'
        smtp_port = 587
        smtp_username = 'smtp username'
        smtp_password = 'password'
        smtp_conn = smtplib.SMTP(smtp_server, smtp_port)
        smtp_conn.starttls()
        smtp_conn.login(smtp_username, smtp_password)
        smtp_conn.sendmail(msg['From'], msg['To'], msg.as_string())
        smtp_conn.quit()

        # Delete the saved attendance sheet from the disk
        os.remove(filename)