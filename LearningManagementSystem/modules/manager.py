import openpyxl
from modules.trainer import Trainer


class Manager(Trainer):
    def __init__(self, trainer_id, full_name, email, phone_number, managed_courses=[], managed_trainers=[]):
        super().__init__(trainer_id, full_name, email, phone_number)
        self._managed_courses = managed_courses
        self._managed_trainers = managed_trainers

    def get_managed_courses(self):
        return self._managed_courses

    def set_managed_courses(self, managed_courses):
        self._managed_courses = managed_courses

    def get_managed_trainers(self):
        return self._managed_trainers

    def set_managed_trainers(self, managed_trainers):
        self._managed_trainers = managed_trainers

    def save_to_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['Managers']
            last_row = ws.max_row
            ws.cell(row=last_row + 1, column=1, value=self._trainer_id)
            ws.cell(row=last_row + 1, column=2, value=self._full_name)
            ws.cell(row=last_row + 1, column=3, value=self._email)
            ws.cell(row=last_row + 1, column=4, value=self._phone_number)
            ws.cell(row=last_row + 1, column=5, value=','.join(self._managed_courses))
            ws.cell(row=last_row + 1, column=6, value=','.join(self._managed_trainers))
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error saving manager to Excel file: {e}")

    def update_in_excel(self, new_full_name=None, new_email=None, new_phone_number=None, new_managed_courses=None,
                        new_managed_trainers=None):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['Managers']
            for row in ws.iter_rows(min_row=2, max_col=6):
                if row[0].value == self._trainer_id:
                    if new_full_name is not None:
                        row[1].value = new_full_name
                    if new_email is not None:
                        row[2].value = new_email
                    if new_phone_number is not None:
                        row[3].value = new_phone_number
                    if new_managed_courses is not None:
                        row[4].value = ','.join(new_managed_courses)
                    if new_managed_trainers is not None:
                        row[5].value = ','.join(new_managed_trainers)
                    break
            else:
                print("Manager not found.")
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error updating manager in Excel file: {e}")

    def delete_from_excel(self):
        try:
            wb = openpyxl.load_workbook('MasterRecord.xlsx')
            ws = wb['Managers']
            for row in ws.iter_rows(min_row=2, max_col=6):
                if row[0].value == self._trainer_id:
                    ws.delete_rows(row[0].row, 1)
                    break
            else:
                print("Manager not found.")
            wb.save('MasterRecord.xlsx')
        except Exception as e:
            print(f"Error deleting manager from Excel file: {e}")
