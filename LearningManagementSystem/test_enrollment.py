import unittest
import openpyxl

from modules.enrollment import Enrollment


class TestEnrollment(unittest.TestCase):
    def test_save_to_excel(self):
        e = Enrollment('C001', 'T001')
        e.save_to_excel()
        # Check if the enrollment record was added to the Excel file
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['Enrollments']
        course_id = ws.cell(row=ws.max_row, column=1).value
        trainee_id = ws.cell(row=ws.max_row, column=2).value
        self.assertEqual(course_id, 'C001')
        self.assertEqual(trainee_id, 'T001')

    def test_delete_from_excel(self):
        e = Enrollment('C001', 'T002')
        e.save_to_excel()
        e.delete_from_excel()
        # Check if the enrollment record was deleted from the Excel file
        wb = openpyxl.load_workbook('MasterRecord.xlsx')
        ws = wb['Enrollments']
        trainee_id = None
        for row in ws.iter_rows(min_row=2, max_col=2):
            if row[0].value == 'C001' and row[1].value == 'T002':
                trainee_id = row[1].value
                break
        self.assertIsNone(trainee_id)


if __name__ == '__main__':
    unittest.main()
